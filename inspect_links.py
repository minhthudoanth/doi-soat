import urllib.request
import openpyxl
import io
import sys

if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

url = 'https://docs.google.com/spreadsheets/d/1ve3IoARSwWmAv_Gz6O8wAJ1YXuNWJNEpQVR7R3iPm-s/export?format=xlsx'
req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
data = urllib.request.urlopen(req).read()
wb = openpyxl.load_workbook(io.BytesIO(data))
s = wb.active

for r in range(1, s.max_row + 1):
    c1 = s.cell(r, 1)
    c2 = s.cell(r, 2)
    h1 = c1.hyperlink.target if c1.hyperlink else None
    h2 = c2.hyperlink.target if c2.hyperlink else None
    val1 = str(c1.value or '').strip()
    val2 = str(c2.value or '').strip()
    if val1 or val2 or h1 or h2:
        print(f"Row {r:02d}: [{val1}] (link: {h1}) | [{val2}] (link: {h2})")
